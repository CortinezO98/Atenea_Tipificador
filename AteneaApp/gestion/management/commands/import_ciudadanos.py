import sys, unicodedata, csv, re
from pathlib import Path
from typing import Dict, Optional, List
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook  

from gestion.models import (
    Ciudadano, Pais, TipoIdentificacion, CalidadComunicacion,
    Discapacidad, Estrato, Genero, GrupoEtnico, GrupoPoblacional,
    Localidad, NivelEducativo, OrientacionSexual, RangoEdad, Sexo,
    TieneDiscapacidad,
)

# helpers 
def norm(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = s.strip()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s.lower()

def build_cache(model, field="nombre", extra_map: Dict[str, int]=None) -> Dict[str, int]:
    m = {}
    for obj in model.objects.values("id", field):
        m[norm(obj[field])] = obj["id"]
    if extra_map:
        m.update({norm(k): v for k, v in extra_map.items()})
    return m

def pick_id(cache: Dict[str, int], label: Optional[str]) -> Optional[int]:
    key = norm(label or "")
    return cache.get(key)

# patch mínimo sugerido 
def limit(s: str, n: int) -> str:
    s = s or ""
    return s[:n]

def clean_phone(raw: str) -> str:
    raw = raw or ""
    raw = raw.strip()
    plus = raw.startswith('+')
    digits = re.sub(r'\D+', '', raw)
    if plus:
        digits = '+' + digits
    return limit(digits, 20)

class Command(BaseCommand):
    help = "Importa ciudadanos desde XLSX, CSV o TSV. Si ya existe, lo omite. Campos vacíos se guardan vacíos. Genera CSV de omitidos."

    def add_arguments(self, parser):
        parser.add_argument("ruta", type=str, help="Ruta del archivo .xlsx/.csv/.tsv")
        parser.add_argument("--sep", default="\t", help="Separador (para CSV/TSV, por defecto TAB)")
        parser.add_argument("--batch", type=int, default=5000, help="Tamaño de lote")
        parser.add_argument("--dry", action="store_true", help="Solo valida sin escribir")

    def handle(self, *args, **o):
        ruta = Path(o["ruta"])
        sep = o["sep"]
        batch_size = o["batch"]
        dry = o["dry"]

        if not ruta.exists():
            self.stderr.write(f"Archivo no existe: {ruta}")
            sys.exit(2)

        ext = ruta.suffix.lower()
        if ext not in [".xlsx", ".csv", ".tsv"]:
            self.stderr.write("Formato no soportado. Usa .xlsx, .csv o .tsv")
            sys.exit(3)

        # ---- caches de catálogos ----
        pais_cache   = build_cache(Pais)
        tipoid_cache = build_cache(TipoIdentificacion)
        calidad_cache= build_cache(CalidadComunicacion)
        disca_cache  = build_cache(Discapacidad)
        estrato_cache= build_cache(Estrato)
        genero_cache = build_cache(Genero)
        etnico_cache = build_cache(GrupoEtnico)
        poblac_cache = build_cache(GrupoPoblacional)
        local_cache  = build_cache(Localidad)
        nivel_cache  = build_cache(NivelEducativo)
        orient_cache = build_cache(OrientacionSexual)
        rango_cache  = build_cache(RangoEdad)
        sexo_cache   = build_cache(Sexo)
        tiene_disc_cache = build_cache(TieneDiscapacidad, extra_map={"si":1, "sí":1, "no":2})

        total = nuevos = saltados = omitidos = 0
        omitidos_lista: List[str] = []
        saltados_largos: List[str] = []   

        # lector universal 
        def iter_filas():
            """Lee XLSX o CSV/TSV y devuelve dicts."""
            if ext == ".xlsx":
                wb = load_workbook(ruta, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    yield dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
            else:
                with ruta.open("r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f, delimiter=sep)
                    for row in reader:
                        yield {k.strip(): (v.strip() if v else "") for k, v in row.items()}

        def parse_row(r: dict) -> Optional[Ciudadano]:
            nonlocal saltados
            # límites y saneos 
            numero = (r.get("Numero Documento") or "").strip()
            if not numero:
                saltados += 1
                return None
            if len(numero) > 20:
                saltados += 1
                saltados_largos.append(numero)
                return None

            tipo_doc_id = pick_id(tipoid_cache, r.get("TipoDocumento"))
            if not tipo_doc_id:
                saltados += 1
                return None

            nombres   = limit((r.get("nombres") or "").strip(), 255)
            apellidos = limit((r.get("apellidos") or "").strip(), 255)
            telefono  = clean_phone(r.get("Celular") or "")
            correo    = limit((r.get("correo_electronico") or "").strip(), 254)
            municipio = limit((r.get("Municipio") or "").strip(), 100)

            sexo_id    = pick_id(sexo_cache, r.get("Sexo"))
            genero_id  = pick_id(genero_cache, r.get("genero"))
            orient_id  = pick_id(orient_cache, r.get("orientación_sexual"))
            rango_id   = pick_id(rango_cache, r.get("rango_edad"))
            nivel_id   = pick_id(nivel_cache, r.get("NivelEducativo"))
            etnico_id  = pick_id(etnico_cache, r.get("g_etnico"))
            poblac_id  = pick_id(poblac_cache, r.get("Poblacion"))
            estrato_id = pick_id(estrato_cache, r.get("Estrato"))
            local_id   = pick_id(local_cache, r.get("Localidad"))
            calidad_id = pick_id(calidad_cache, r.get("TeComunicasEn"))
            pais_id    = pick_id(pais_cache, r.get("Pais"))

            disc_label = (r.get("discapacidad") or "").strip()
            td_id = pick_id(tiene_disc_cache, disc_label)
            disca_id = None
            
            if not td_id and pick_id(disca_cache, disc_label):
                td_id = 1
                disca_id = pick_id(disca_cache, disc_label)
            elif td_id == 1:
                disca_id = pick_id(disca_cache, disc_label)

            nombre_completo = limit((f"{nombres} {apellidos}").strip(), 255)

            return Ciudadano(
                numero_identificacion=numero,
                nombre=nombre_completo,
                correo=correo,                 
                telefono=telefono,             
                direccion_residencia="",       
                ciudad=(municipio or None),    
                pais_id=pais_id,
                tipo_identificacion_id=tipo_doc_id,
                calidad_comunicacion_id=calidad_id,
                discapacidad_id=disca_id,
                estrato_id=estrato_id,
                genero_id=genero_id,
                grupo_etnico_id=etnico_id,
                grupo_poblacional_id=poblac_id,
                localidad_id=local_id,
                nivel_educativo_id=nivel_id,
                orientacion_id=orient_id,
                rango_edad_id=rango_id,
                sexo_id=sexo_id,
                tiene_discapacidad_id=td_id,
            )

        # función para procesar lotes 
        def flush(rows):
            nonlocal nuevos, omitidos
            if not rows:
                return
            docs = [(r.get("Numero Documento") or "").strip() for r in rows if (r.get("Numero Documento") or "").strip()]
            existentes = set(
                Ciudadano.objects.filter(numero_identificacion__in=docs)
                .values_list("numero_identificacion", flat=True)
            )

            crear = []
            for r in rows:
                obj = parse_row(r)
                if not obj:
                    continue
                if obj.numero_identificacion in existentes:
                    omitidos += 1
                    omitidos_lista.append(obj.numero_identificacion)
                    continue
                crear.append(obj)

            if not dry and crear:
                with transaction.atomic():
                    Ciudadano.objects.bulk_create(crear, batch_size=len(crear))
                    nuevos += len(crear)

        # ejecutar 
        batch = []
        for row in iter_filas():
            total += 1
            batch.append(row)
            if len(batch) >= batch_size:
                flush(batch)
                batch.clear()
        flush(batch)

        # guardar omitidos en CSV 
        if omitidos_lista:
            out_path = ruta.parent / "ciudadanos_omitidos.csv"
            with out_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["numero_identificacion"])
                for num in omitidos_lista:
                    writer.writerow([num])
            self.stdout.write(self.style.WARNING(
                f"Archivo generado: {out_path} ({len(omitidos_lista)} omitidos)"
            ))

        # guardar saltados por doc largo
        if saltados_largos:
            out2 = ruta.parent / "ciudadanos_saltados_doc_largo.csv"
            with out2.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["numero_identificacion"])
                for num in saltados_largos:
                    w.writerow([num])
            self.stdout.write(self.style.WARNING(
                f"Archivo generado: {out2} ({len(saltados_largos)} saltados por doc > 20)"
            ))

        # resumen final 
        resumen = (
            f"Procesadas: {total:,} | Nuevos: {nuevos:,} | "
            f"Omitidos (ya existen): {omitidos:,} | Saltados: {saltados:,}"
            + (" (dry-run)" if dry else "")
        )
        self.stdout.write(self.style.SUCCESS(resumen))

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.contrib import messages
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.decorators.http import require_GET, require_http_methods
from django.utils.timezone import now, make_aware, localtime
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.urls import reverse
from usuarios.views import ValidarRolUsuario, en_grupo
from usuarios.enums import Roles
from .models import *
from .utils import RegistrarError, get_ciudadano_anonimo
from io import BytesIO
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
import inspect
from django.views.decorators.csrf import csrf_exempt  # puedes dejarlo si lo usas en otros lados
from django.utils.html import format_html
from .forms import build_encuesta_form


def index(request):
    if request.user.is_authenticated:
        if ValidarRolUsuario(request, Roles.ADMINISTRADOR.value):
            return redirect('dashboard_admin')
        elif ValidarRolUsuario(request, Roles.SUPERVISOR.value):
            return redirect('dashboard_supervisor')
        elif ValidarRolUsuario(request, Roles.AGENTE.value):
            return redirect('dashboard_agente')
        messages.warning(request, "Usted no esta autorizado.")
        return redirect('logout')
    else:
        return redirect('login')


def _clean_str(value, max_len=None):
    """
    Normaliza strings desde request:
    - Convierte None a ''
    - Hace strip
    - (Opcional) corta a max_len
    """
    v = (value or '').strip()
    if max_len is not None:
        return v[:max_len]
    return v


CIUDADANO_FIELDS = (
    'tipo_identificacion_id',
    'numero_identificacion',
    'nombre',
    'correo',
    'telefono',
    'direccion_residencia',
    'pais_id',
    'ciudad',
    'sexo_id',
    'genero_id',
    'orientacion_id',
    'tiene_discapacidad_id',
    'discapacidad_id',
    'rango_edad_id',
    'nivel_educativo_id',
    'grupo_etnico_id',
    'grupo_poblacional_id',
    'estrato_id',
    'localidad_id',
    'calidad_comunicacion_id',
)


@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value, Roles.AGENTE.value])
def crear_evaluacion(request):
    """
    Crea Evaluación:
      - Captura/actualiza Ciudadano (con caracterización por IDs si NO es anónimo)
      - Guarda categoría final (árbol N1..N6)
      - Guarda tipo_canal_id
      - Crea encuesta asociada
    """
    is_admin = ValidarRolUsuario(request, Roles.ADMINISTRADOR.value)
    is_supervisor = ValidarRolUsuario(request, Roles.SUPERVISOR.value)
    is_agent = ValidarRolUsuario(request, Roles.AGENTE.value)
    if not (is_admin or is_supervisor or is_agent):
        return HttpResponseForbidden("No autorizado")

    if request.method == 'POST':
        try:
            with transaction.atomic():
                numero_doc = (request.POST.get('numero_identificacion') or '').strip()
                es_anonimo = (
                    request.POST.get('es_anonimo') == '1'
                    or numero_doc in ('9999', '9999999')
                )

                if es_anonimo:
                    ciudadano = get_ciudadano_anonimo()
                else:
                    # IMPORTANTE: validación adicional para evitar que un usuario
                    # manipule el id de ciudadano y actualice datos de otro.
                    cid_raw = request.POST.get('cuidadano_id')
                    tipo_identificacion_id = request.POST['tipo_identificacion']
                    numero_identificacion = _clean_str(
                        request.POST['numero_identificacion'], 20
                    )
                    nombre = _clean_str(request.POST['nombre'], 255)
                    correo = _clean_str(request.POST.get('correo', ''), 254)
                    telefono = _clean_str(request.POST.get('telefono', ''), 20)
                    direccion_residencia = _clean_str(
                        request.POST.get('direccion_residencia', ''), 255
                    )
                    pais_id = request.POST.get('pais') or None
                    ciudad = _clean_str(request.POST.get('ciudad', ''), 255)
                    sexo_id = request.POST.get('sexo') or None
                    genero_id = request.POST.get('genero') or None
                    orientacion_id = request.POST.get('orientacion') or None
                    tiene_discapacidad_id = request.POST.get('tiene_discapacidad') or None
                    if request.POST.get('tiene_discapacidad') == '1':
                        discapacidad_id = request.POST.get('discapacidad') or None
                    else:
                        discapacidad_id = None
                    rango_edad_id = request.POST.get('rango_edad') or None
                    nivel_educativo_id = request.POST.get('nivel_educativo') or None
                    grupo_etnico_id = request.POST.get('grupo_etnico') or None
                    grupo_poblacional_id = request.POST.get('grupo_poblacional') or None
                    estrato_id = request.POST.get('estrato') or None
                    localidad_id = request.POST.get('localidad') or None
                    calidad_comunicacion_id = (
                        request.POST.get('calidad_comunicacion') or None
                    )

                    ciudadano = None
                    if cid_raw and str(cid_raw).isdigit():
                        # Object-Level Authorization defensivo:
                        # se asegura que el id coincida con el documento y tipo
                        ciudadano = get_object_or_404(
                            Ciudadano,
                            id=int(cid_raw),
                            numero_identificacion=numero_identificacion,
                            tipo_identificacion_id=tipo_identificacion_id,
                        )

                    if ciudadano:
                        ciudadano.tipo_identificacion_id = tipo_identificacion_id
                        ciudadano.numero_identificacion = numero_identificacion
                        ciudadano.nombre = nombre
                        ciudadano.correo = correo
                        ciudadano.telefono = telefono
                        ciudadano.direccion_residencia = direccion_residencia
                        ciudadano.pais_id = pais_id
                        ciudadano.ciudad = ciudad
                        ciudadano.sexo_id = sexo_id
                        ciudadano.genero_id = genero_id
                        ciudadano.orientacion_id = orientacion_id
                        ciudadano.tiene_discapacidad_id = tiene_discapacidad_id
                        ciudadano.discapacidad_id = discapacidad_id
                        ciudadano.rango_edad_id = rango_edad_id
                        ciudadano.nivel_educativo_id = nivel_educativo_id
                        ciudadano.grupo_etnico_id = grupo_etnico_id
                        ciudadano.grupo_poblacional_id = grupo_poblacional_id
                        ciudadano.estrato_id = estrato_id
                        ciudadano.localidad_id = localidad_id
                        ciudadano.calidad_comunicacion_id = (
                            calidad_comunicacion_id
                        )
                        ciudadano.save()
                    else:
                        # Si el id no es válido o no coincide, se crea un nuevo registro.
                        ciudadano = Ciudadano.objects.create(
                            tipo_identificacion_id=tipo_identificacion_id,
                            numero_identificacion=numero_identificacion,
                            nombre=nombre,
                            correo=correo,
                            telefono=telefono,
                            direccion_residencia=direccion_residencia,
                            pais_id=pais_id,
                            ciudad=ciudad,
                            sexo_id=sexo_id,
                            genero_id=genero_id,
                            orientacion_id=orientacion_id,
                            tiene_discapacidad_id=tiene_discapacidad_id,
                            discapacidad_id=discapacidad_id,
                            rango_edad_id=rango_edad_id,
                            nivel_educativo_id=nivel_educativo_id,
                            grupo_etnico_id=grupo_etnico_id,
                            grupo_poblacional_id=grupo_poblacional_id,
                            estrato_id=estrato_id,
                            localidad_id=localidad_id,
                            calidad_comunicacion_id=calidad_comunicacion_id,
                        )

                categoria_final_id = None
                for key in ('nivel6', 'nivel5', 'nivel4', 'nivel3', 'nivel2', 'nivel1'):
                    val = request.POST.get(key)
                    if val and val != "0":
                        categoria_final_id = val
                        break
                if not categoria_final_id:
                    raise ValueError(
                        "Debes seleccionar al menos una categoría (N1..N6)."
                    )

                tel_inconser = _clean_str(
                    request.POST.get('telefono_inconser')
                    or request.POST.get('telefono_inconcer')
                    or '',
                    20,
                )

                evaluacion = Evaluacion.objects.create(
                    conversacion_id=request.POST['conversacion_id'],
                    observacion=request.POST['observacion'],
                    ciudadano=ciudadano,
                    user=request.user,
                    categoria_id=categoria_final_id,
                    tipo_canal_id=(request.POST.get('tipo_canal') or None),
                    segmento=None,
                    segmento_ii=None,
                    segmento_iii=None,
                    segmento_iv=None,
                    segmento_v=None,
                    segmento_vi=None,
                    tipificacion=None,
                    es_anonimo=es_anonimo,
                    contacto_correo=(
                        request.POST.get('correo', '') if es_anonimo else None
                    ),
                    contacto_telefono=(
                        request.POST.get('telefono', '') if es_anonimo else None
                    ),
                    contacto_telefono_inconcer=tel_inconser,
                )

                token = get_random_string(24)
                expira = now() + timedelta(hours=24)
                encuesta = Encuesta.objects.create(
                    evaluacion=evaluacion,
                    agente=request.user,
                    idInteraccion=request.POST.get('conversacion_id', ''),
                    nombreAgente=(
                        request.user.get_full_name() or request.user.username
                    ),
                    token=token,
                    fechaExpiracionLink=expira,
                    fecha_creacion=now(),
                )
                encuesta_url = request.build_absolute_uri(
                    reverse('encuesta_publica', kwargs={'token': token})
                )

                debug_info = {
                    'nivel1': request.POST.get('nivel1'),
                    'nivel2': request.POST.get('nivel2'),
                    'nivel3': request.POST.get('nivel3'),
                    'nivel4': request.POST.get('nivel4'),
                    'nivel5': request.POST.get('nivel5'),
                    'nivel6': request.POST.get('nivel6'),
                    'categoria_final_id': categoria_final_id,
                    'tipo_canal': request.POST.get('tipo_canal'),
                    'es_anonimo': request.POST.get('es_anonimo'),
                    'telefono_inconcer': tel_inconser,
                    'encuesta_token': encuesta.token,
                }

                if settings.DEBUG:
                    print(f"DEBUG - Crear evaluación: {debug_info}")

                messages.success(
                    request,
                    format_html(
                        '✅ <strong>Evaluación guardada exitosamente</strong><br>'
                        '<div class="text-center mt-2">'
                        '<span class="text-muted small ms-2 link-copied-feedback" style="display:none;">'
                        '<i class="bi bi-check-circle text-success me-1"></i>¡Copiado!'
                        '</span>'
                        '</div>'
                        '<div class="mt-2 p-2 bg-light rounded">'
                        '<code class="text-break small d-block">{}</code>'
                        '</div>'
                        '<small class="text-muted d-block mt-1 text-center">'
                        'Comparte este enlace para acceder a la evaluación'
                        '</small>',
                        encuesta_url,
                    ),
                )

        except Exception as e:
            RegistrarError(inspect.currentframe().f_code.co_name, str(e), request)
            messages.error(
                request,
                f"Ocurrió un error al guardar la evaluación: {str(e)}",
            )
        return redirect('index')

    tiposIdentificacion = TipoIdentificacion.objects.all()
    paises = Pais.objects.all()
    hay_nivel1 = Categoria.objects.filter(nivel=1).exists()

    if (not tiposIdentificacion.exists()) or (not paises.exists()) or (not hay_nivel1):
        messages.warning(
            request,
            "Faltan datos mínimos (Tipos de identificación, Países o Niveles). "
            "Ejecuta los seeders mínimos y vuelve a intentar.",
        )
        return redirect('index')

    return render(
        request,
        'usuarios/evaluaciones/crear_evaluacion.html',
        {
            'tiposIdentificacion': tiposIdentificacion,
            'paises': paises,
            'is_supervisor': (
                ValidarRolUsuario(request, Roles.SUPERVISOR.value)
                or ValidarRolUsuario(request, Roles.ADMINISTRADOR.value)
            ),
            'is_agente': ValidarRolUsuario(request, Roles.AGENTE.value),
            'is_admin': ValidarRolUsuario(request, Roles.ADMINISTRADOR.value),
            'numero_anonimo': '9999999',
            'sexos': Sexo.objects.all(),
            'generos': Genero.objects.all(),
            'orientaciones': OrientacionSexual.objects.all(),
            'tiene_discapacidades': TieneDiscapacidad.objects.all(),
            'discapacidades': Discapacidad.objects.all(),
            'rangos_edad': RangoEdad.objects.all(),
            'niveles_educativos': NivelEducativo.objects.all(),
            'grupos_etnicos': GrupoEtnico.objects.all(),
            'grupos_poblacionales': GrupoPoblacional.objects.all(),
            'estratos': Estrato.objects.all(),
            'localidades': Localidad.objects.all(),
            'calidades': CalidadComunicacion.objects.all(),
            'tipos_canal': TipoCanal.objects.all(),
        },
    )


@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value, Roles.AGENTE.value])
def buscar_tipificacion(request):
    """
    Busca Evaluacion(es) por número de identificación del ciudadano,
    correo/teléfono de contacto guardados en la Evaluación (anónimo),
    y correo/telefoneo del Ciudadano (no anónimo / legacy).
    """
    query = (request.GET.get('q') or '').strip()
    evaluaciones = Evaluacion.objects.none()
    paginator = None

    if query:
        qs = (
            Evaluacion.objects
            .select_related(
                'ciudadano__tipo_identificacion',
                'ciudadano__pais',
                'categoria__tipificacion',
                'categoria__categoria_padre',
                'user',
                'tipo_canal',
                'segmento',
                'segmento_ii',
                'segmento_iii',
                'segmento_iv',
                'segmento_v',
                'segmento_vi',
                'tipificacion'
            )
            .filter(
                Q(ciudadano__numero_identificacion__icontains=query) |
                Q(contacto_telefono__icontains=query) |
                Q(contacto_correo__icontains=query) |
                Q(contacto_telefono_inconcer__icontains=query) |
                Q(ciudadano__telefono__icontains=query) |
                Q(ciudadano__correo__icontains=query)
            )
            .order_by('-fecha')
        )

        is_admin = ValidarRolUsuario(request, Roles.ADMINISTRADOR.value)
        is_supervisor = ValidarRolUsuario(request, Roles.SUPERVISOR.value)
        is_agent = ValidarRolUsuario(request, Roles.AGENTE.value)

        # Object-Level Authorization: el agente solo ve sus propias evaluaciones
        if is_agent and not (is_admin or is_supervisor):
            qs = qs.filter(user=request.user)

        paginator = Paginator(qs, 10)
        page_number = request.GET.get('page')
        if not page_number or not str(page_number).isdigit():
            page_number = 1
        evaluaciones = paginator.get_page(page_number)

    return render(request, 'usuarios/evaluaciones/buscar_tipificacion.html', {
        'evaluaciones': evaluaciones,
        'paginator':    paginator,
        'page_obj':     evaluaciones,
        'is_paginated': evaluaciones.has_other_pages() if evaluaciones else False,
        'query':        query,
    })


def get_quick_range(quick: str):
    """
    quick ∈ {'hoy', 'ayer', '7d'}
    Retorna (start, end, etiqueta) timezone-aware.
    """
    tz_now = now()
    if quick == 'hoy':
        start = tz_now.replace(hour=0, minute=0, second=0, microsecond=0)
        end   = tz_now
        return start, end, 'hoy'
    if quick == 'ayer':
        ayer  = tz_now - timedelta(days=1)
        start = ayer.replace(hour=0, minute=0, second=0, microsecond=0)
        end   = start + timedelta(days=1)
        return start, end, 'ayer'
    if quick == '7d':
        start = tz_now - timedelta(days=7)
        end   = tz_now
        return start, end, '7d'
    return None, None, None


@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value, Roles.AGENTE.value])
def reportes_view(request):
    quick = request.GET.get('quick')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    usuario_id = request.GET.get('usuario')
    canal_id = request.GET.get('canal')

    is_admin = ValidarRolUsuario(request, Roles.ADMINISTRADOR.value)
    is_supervisor = ValidarRolUsuario(request, Roles.SUPERVISOR.value)
    is_agent = ValidarRolUsuario(request, Roles.AGENTE.value)
    if not (is_admin or is_supervisor or is_agent):
        return HttpResponseForbidden("No autorizado")

    qs = (
        Evaluacion.objects
        .select_related(
            'ciudadano__tipo_identificacion',
            'ciudadano__pais',
            'categoria__tipificacion',
            'categoria__categoria_padre',
            'user',
            'tipo_canal',
            'segmento',
            'segmento_ii',
            'segmento_iii',
            'segmento_iv',
            'segmento_v',
            'segmento_vi',
            'tipificacion'
        )
        .order_by('-fecha')
    )

    # El agente solo ve sus propias evaluaciones (Object-Level Authorization)
    if is_agent and not (is_admin or is_supervisor):
        qs = qs.filter(user=request.user)

    if quick in ('hoy', 'ayer', '7d'):
        start, end, _ = get_quick_range(quick)
        if quick == 'ayer':
            qs = qs.filter(fecha__gte=start, fecha__lt=end)
        else:
            qs = qs.filter(fecha__range=(start, end))
    elif fecha_inicio and fecha_fin:
        try:
            fmt = '%Y-%m-%d'
            fi = make_aware(datetime.strptime(fecha_inicio, fmt))
            ff = make_aware(datetime.strptime(fecha_fin,    fmt))
            qs = qs.filter(fecha__range=(fi, ff))
        except ValueError:
            pass
    else:
        start, end, _ = get_quick_range('hoy')
        qs = qs.filter(fecha__range=(start, end))
        quick = 'hoy'

    if usuario_id and str(usuario_id).isdigit():
        qs = qs.filter(user_id=int(usuario_id))
    if canal_id and str(canal_id).isdigit():
        qs = qs.filter(tipo_canal_id=int(canal_id))

    # Usuarios visibles en filtros
    if is_agent and not (is_admin or is_supervisor):
        usuarios = User.objects.filter(id=request.user.id).distinct()
    else:
        usuarios = User.objects.filter(
            groups__name__in=['Agente', 'Supervisor', 'Administrador']
        ).distinct()

    canales = TipoCanal.objects.all()

    # Base para totales (respetando las mismas restricciones de acceso)
    base = Evaluacion.objects.all()
    if is_agent and not (is_admin or is_supervisor):
        base = base.filter(user=request.user)

    if usuario_id and str(usuario_id).isdigit():
        base = base.filter(user_id=int(usuario_id))
    if canal_id and str(canal_id).isdigit():
        base = base.filter(tipo_canal_id=int(canal_id))

    hoy_start, hoy_end, _ = get_quick_range('hoy')
    ayer_start, ayer_end, _ = get_quick_range('ayer')
    semana_start, semana_end, _ = get_quick_range('7d')

    total_hoy = base.filter(fecha__range=(hoy_start, hoy_end)).count()
    total_ayer = base.filter(fecha__gte=ayer_start, fecha__lt=ayer_end).count()
    total_7d = base.filter(fecha__range=(semana_start, semana_end)).count()

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    return render(request, 'usuarios/reportes.html', {
        'reportes': page_obj,
        'paginator': paginator,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'hoy': now().date().isoformat(),
        'quick': quick,
        'usuarios': usuarios,
        'canales': canales,
        'total_hoy': total_hoy,
        'total_ayer': total_ayer,
        'total_7d': total_7d,
    })


def xl_safe(v):
    if v is None:
        return ''
    return ILLEGAL_CHARACTERS_RE.sub('', str(v))


# Límite de filas para exportación para evitar DoS por bucle excesivo
MAX_EXPORT_ROWS = getattr(settings, 'MAX_EXPORT_ROWS', 50000)


@require_GET
@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value])
def exportar_excel(request):
    if not (
        ValidarRolUsuario(request, Roles.ADMINISTRADOR.value)
        or ValidarRolUsuario(request, Roles.SUPERVISOR.value)
    ):
        return HttpResponseForbidden("No autorizado")

    quick        = request.GET.get('quick')
    hoy_flag     = request.GET.get('hoy') == '1'
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin    = request.GET.get('fecha_fin')
    usuario      = request.GET.get('usuario')
    canal        = request.GET.get('canal')

    qs = (
        Evaluacion.objects
        .select_related(
            'ciudadano__tipo_identificacion',
            'ciudadano__pais',
            'ciudadano__sexo',
            'ciudadano__genero',
            'ciudadano__orientacion',
            'ciudadano__tiene_discapacidad',
            'ciudadano__discapacidad',
            'ciudadano__rango_edad',
            'ciudadano__nivel_educativo',
            'ciudadano__grupo_etnico',
            'ciudadano__grupo_poblacional',
            'ciudadano__estrato',
            'ciudadano__localidad',
            'ciudadano__calidad_comunicacion',
            'categoria__tipificacion',
            'categoria__categoria_padre',
            'user',
            'tipo_canal',
            'segmento', 'segmento_ii', 'segmento_iii',
            'segmento_iv', 'segmento_v', 'segmento_vi', 'tipificacion',
        )
        .order_by('-fecha')
    )

    # Filtros de fechas
    if quick in ('hoy', 'ayer', '7d') or hoy_flag:
        use_quick = quick or 'hoy'
        start, end, _ = get_quick_range(use_quick)
        if use_quick == 'ayer':
            qs = qs.filter(fecha__gte=start, fecha__lt=end)
        else:
            qs = qs.filter(fecha__range=(start, end))
    elif fecha_inicio and fecha_fin:
        try:
            fmt = '%Y-%m-%d'
            fi = make_aware(datetime.strptime(fecha_inicio, fmt))
            ff = make_aware(datetime.strptime(fecha_fin,    fmt))
            qs = qs.filter(fecha__range=(fi, ff))
        except ValueError:
            pass
    else:
        # Si no hay filtros específicos de usuario/canal, se limita por defecto a 'hoy'
        if not (usuario and str(usuario).isdigit()) and not (canal and str(canal).isdigit()):
            start, end, _ = get_quick_range('hoy')
            qs = qs.filter(fecha__range=(start, end))

    if usuario and str(usuario).isdigit():
        qs = qs.filter(user_id=int(usuario))
    if canal and str(canal).isdigit():
        qs = qs.filter(tipo_canal_id=int(canal))

    # MITIGACIÓN Unchecked_Input_for_Loop_Condition:
    # Limitar máximo de filas exportadas para evitar loops enormes controlados por parámetros.
    qs = qs[:MAX_EXPORT_ROWS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = [
        'Fecha', 'Tipo Documento', 'Número ID', 'Nombre',
        'Correo', 'Teléfono',
        'País', 'Ciudad', 'Dirección',
        'Sexo','Género','Orientación Sexual','Tiene Discapacidad','Discapacidad',
        'Rango de Edad','Nivel Educativo','Grupo Étnico','Grupo Poblacional',
        'Estrato','Localidad','Calidad Comunicación',
        'Correo contacto', 'Teléfono contacto', 'Teléfono Inconcer',
        'ID Conversación',
        'Nivel 1', 'Nivel 2', 'Nivel 3', 'Nivel 4', 'Nivel 5', 'Nivel 6', 'Nivel Final',
        'Encuesta estado', 'Encuesta respondida en', 'Encuesta token', 'Encuesta URL',
        'Encuesta promedio (escala 1-5)', 'Encuesta respuestas (texto)',
        'Tipo de Canal',
        'Observación',
        'Usuario'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    from .models import Encuesta, RespuestaEncuesta

    for row_idx, ev in enumerate(qs.iterator(), start=2):
        ciu = ev.ciudadano

        correo_base     = ciu.correo or ''
        telf_base       = ciu.telefono or ''
        correo_contacto = ev.contacto_correo or ''
        telf_contacto   = ev.contacto_telefono or ''
        telf_inconser   = ev.contacto_telefono_inconcer or ''

        n1 = n2 = n3 = n4 = n5 = n6 = ''
        nivel_final = 'Sin categoría'
        if ev.categoria:
            cur = ev.categoria
            niveles = {}
            while cur:
                niveles[cur.nivel] = cur.nombre
                cur = cur.categoria_padre
            n1 = niveles.get(1, '')
            n2 = niveles.get(2, '')
            n3 = niveles.get(3, '')
            n4 = niveles.get(4, '')
            n5 = niveles.get(5, '')
            n6 = niveles.get(6, '')
            for k in (6, 5, 4, 3, 2, 1):
                if niveles.get(k):
                    nivel_final = f"Nivel {k}"
                    break

        encuesta = (
            Encuesta.objects
            .filter(evaluacion=ev)
            .order_by('-fecha_creacion')
            .first()
        )

        enc_estado = 'Pendiente'
        enc_resp_en = ''
        enc_token = ''
        enc_url = ''
        enc_prom_escala = ''
        enc_respuestas_txt = ''

        if encuesta:
            try:
                if getattr(encuesta, 'cerrada', False):
                    enc_estado = 'Cerrada'
                    if getattr(encuesta, 'expirada', False):
                        enc_estado = 'Expirada'
                else:
                    has_resp = RespuestaEncuesta.objects.filter(encuesta=encuesta).exists()
                    enc_estado = 'Respondida' if has_resp else 'Pendiente'
            except Exception:
                has_resp = RespuestaEncuesta.objects.filter(encuesta=encuesta).exists()
                enc_estado = 'Respondida' if has_resp else 'Pendiente'

            if hasattr(encuesta, 'respondida_en') and getattr(encuesta, 'respondida_en'):
                enc_resp_en = localtime(encuesta.respondida_en).strftime('%Y-%m-%d %H:%M')

            enc_token = encuesta.token or ''
            try:
                enc_url = request.build_absolute_uri(
                    reverse('encuesta_publica', kwargs={'token': encuesta.token})
                )
            except Exception:
                enc_url = ''

            respuestas = (
                RespuestaEncuesta.objects
                .filter(encuesta=encuesta)
                .select_related('pregunta')
                .order_by('pregunta_id')
            )

            escala_vals = []
            txt_pairs = []
            for r in respuestas:
                ptxt = getattr(r.pregunta, 'texto', f'Pregunta {r.pregunta_id}') if hasattr(r, 'pregunta') else f'Pregunta {r.pregunta_id}'
                txt_pairs.append(f"{ptxt}: {r.valor}")
                try:
                    if getattr(r.pregunta, 'tipo', '') == 'escala' and str(r.valor).isdigit():
                        escala_vals.append(int(r.valor))
                except Exception:
                    pass

            if escala_vals:
                enc_prom_escala = round(sum(escala_vals) / len(escala_vals), 2)
            enc_respuestas_txt = " | ".join(txt_pairs) if txt_pairs else ''

        data = [
            localtime(ev.fecha).strftime('%Y-%m-%d %H:%M'),
            ciu.tipo_identificacion.nombre,
            ciu.numero_identificacion,
            ciu.nombre,
            ciu.correo or '',
            ciu.telefono or '',
            ciu.pais.nombre if ciu.pais else '',
            ciu.ciudad or '',
            ciu.direccion_residencia or '',
            getattr(ciu.sexo, 'nombre', ''),
            getattr(ciu.genero, 'nombre', ''),
            getattr(ciu.orientacion, 'nombre', ''),
            getattr(ciu.tiene_discapacidad, 'nombre', ''),
            (getattr(ciu.discapacidad, 'nombre', '') if ciu.tiene_discapacidad_id == 1 else ''),
            getattr(ciu.rango_edad, 'nombre', ''),
            getattr(ciu.nivel_educativo, 'nombre', ''),
            getattr(ciu.grupo_etnico, 'nombre', ''),
            getattr(ciu.grupo_poblacional, 'nombre', ''),
            getattr(ciu.estrato, 'nombre', ''),
            getattr(ciu.localidad, 'nombre', ''),
            getattr(ciu.calidad_comunicacion, 'nombre', ''),
            ev.contacto_correo or '',
            ev.contacto_telefono or '',
            ev.contacto_telefono_inconcer or '',
            ev.conversacion_id,
            n1, n2, n3, n4, n5, n6, nivel_final,
            enc_estado, enc_resp_en, enc_token, enc_url,
            enc_prom_escala, enc_respuestas_txt,
            getattr(ev.tipo_canal, 'nombre', ''),
            ev.observacion,
            ev.user.username
        ]
        for col, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col, value=xl_safe(value))

    col_widths = [len(h) + 2 for h in headers]
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for i, cell in enumerate(r):
            val = '' if cell.value is None else str(cell.value)
            w = len(val) + 2
            if w > col_widths[i]:
                col_widths[i] = min(w, 60)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, w)

    buffer = BytesIO()
    try:
        wb.save(buffer)
    except Exception as e:
        RegistrarError('exportar_excel', str(e), request)
        messages.error(request, 'No se pudo generar el Excel. Soporte ha sido notificado.')
        return redirect('reportes_view')

    buffer.seek(0)
    resp = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="reportes_niveles_encuesta.xlsx"'
    return resp


@require_http_methods(["GET", "POST"])
def encuesta_publica(request, token):
    """
    Encuesta pública:
    - Se eliminó @csrf_exempt para que quede protegida por el middleware CSRF.
    - Asegúrate de tener {% csrf_token %} en el formulario HTML de la plantilla.
    """
    # Validación defensiva del token (tamano/charset) para evitar inputs extraños.
    if not isinstance(token, str) or len(token) != 24 or not token.isalnum():
        return HttpResponseBadRequest("Token inválido")

    encuesta = get_object_or_404(Encuesta, token=token)
    if encuesta.cerrada:
        if encuesta.expirada:
            return render(request, 'usuarios/encuestas/expirada.html', {'encuesta': encuesta})
        else:
            return render(request, 'usuarios/encuestas/completada.html', {'encuesta': encuesta})

    preguntas = list(PreguntaEncuesta.objects.all())

    if request.method == 'POST':
        # CSRF ahora es verificado por el middleware (no hay @csrf_exempt)
        for p in preguntas:
            field = f"q_{p.id}"
            val = (request.POST.get(field) or '').strip()
            if p.tipo == 'escala':
                if val not in {'1', '2', '3', '4', '5'}:
                    return render(request, 'usuarios/encuestas/form.html', {
                        'encuesta': encuesta,
                        'preguntas': preguntas,
                        'error': f'Falta responder correctamente: "{p.texto}"'
                    })
            elif p.tipo == 'si_no':
                if val not in {'Si', 'No'}:
                    return render(request, 'usuarios/encuestas/form.html', {
                        'encuesta': encuesta,
                        'preguntas': preguntas,
                        'error': f'Falta responder correctamente: "{p.texto}"'
                    })

            RespuestaEncuesta.objects.update_or_create(
                encuesta=encuesta, pregunta=p, defaults={'valor': val}
            )

        encuesta.respondida_en = timezone.now()
        encuesta.save()

        return render(request, 'usuarios/encuestas/gracias.html', {'encuesta': encuesta})

    respuestas_previas = {
        r.pregunta_id: r.valor for r in RespuestaEncuesta.objects.filter(encuesta=encuesta)
    }

    return render(request, 'usuarios/encuestas/form.html', {
        'encuesta': encuesta,
        'preguntas': preguntas,
        'respuestas_previas': respuestas_previas
    })
